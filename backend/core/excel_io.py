from __future__ import annotations

import re
import uuid
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook

from core.models import Activity, Student, StudentActivity

HEADER_MAP = {
    "学号": "student_id",
    "院系": "department",
    "年级": "grade",
    "姓名": "name",
    "班级": "class_name",
    "学业水平成绩": "academic_score",
    "身心素养活动": "mind_lines",
    "文艺素养活动": "art_lines",
    "劳动素养活动": "labor_lines",
    "创新素养活动": "innovation_lines",
    "素质总分": "_ignore_quality",
    "综合分值": "_ignore_comp",
    "备注": "remark",
}

CATEGORY_BY_FIELD = {
    "mind_lines": Activity.CATEGORY_MIND,
    "art_lines": Activity.CATEGORY_ART,
    "labor_lines": Activity.CATEGORY_LABOR,
    "innovation_lines": Activity.CATEGORY_INNOVATION,
}

LINE_RE = re.compile(r"^(.+?)[（(]([\d.]+)[）)]\s*$")


def _new_activity_id() -> str:
    return f"A{uuid.uuid4().hex[:12].upper()}"


def parse_activity_cell(text: str | None) -> list[tuple[str, Decimal]]:
    if text is None or (isinstance(text, str) and not str(text).strip()):
        return []
    raw = str(text).replace("\r\n", "\n").replace("\r", "\n")
    items: list[tuple[str, Decimal]] = []
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = LINE_RE.match(line)
        if not m:
            raise ValueError(f"无法解析活动行: {line!r}，应为「活动名称（分值）」")
        name = m.group(1).strip()
        try:
            score = Decimal(m.group(2)).quantize(Decimal("0.01"))
        except InvalidOperation as e:
            raise ValueError(f"分值无效: {line!r}") from e
        if score < 0:
            raise ValueError(f"分值不能为负: {line!r}")
        items.append((name, score))
    return items


def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _cell_decimal(val, default=Decimal("0")) -> Decimal:
    if val is None or val == "":
        return default
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as e:
        raise ValueError(f"数字格式错误: {val!r}") from e


REQUIRED_HEADERS = [
    "学号",
    "姓名",
    "院系",
    "年级",
    "班级",
    "学业水平成绩",
    "身心素养活动",
    "文艺素养活动",
    "劳动素养活动",
    "创新素养活动",
    "备注",
]


def parse_import_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 为空")
    headers = [_cell_str(h) for h in rows[0]]
    col_index = {h: i for i, h in enumerate(headers) if h}
    for h in REQUIRED_HEADERS:
        if h not in col_index:
            raise ValueError(f"缺少列: {h}")

    parsed: list[dict] = []
    errors: list[str] = []
    for ridx, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            item: dict = {}
            for zh, key in HEADER_MAP.items():
                if zh not in col_index:
                    continue
                if key.startswith("_ignore"):
                    continue
                cell = row[col_index[zh]] if col_index[zh] < len(row) else None
                if key == "student_id":
                    item["student_id"] = _cell_str(cell)
                elif key == "name":
                    item["name"] = _cell_str(cell)
                elif key in ("department", "grade", "class_name", "remark"):
                    item[key] = _cell_str(cell)
                elif key == "academic_score":
                    item["academic_score"] = _cell_decimal(cell, Decimal("0"))
                elif key in CATEGORY_BY_FIELD:
                    item[key] = parse_activity_cell(_cell_str(cell) if cell else "")
            if not item.get("student_id"):
                errors.append(f"第{ridx}行: 学号为空")
                continue
            if not item.get("name"):
                errors.append(f"第{ridx}行: 姓名为空")
                continue
            parsed.append(item)
        except ValueError as e:
            errors.append(f"第{ridx}行: {e}")
    return parsed, errors


def student_row_to_dict(s: Student) -> dict:
    return {
        "student_id": s.student_id,
        "name": s.name,
        "department": s.department,
        "grade": s.grade,
        "class_name": s.class_name,
        "academic_score": str(s.academic_score),
        "remark": s.remark,
    }


def build_conflict_preview(rows: list[dict]) -> list[dict]:
    out = []
    existing = {
        s.student_id: s
        for s in Student.objects.filter(
            student_id__in=[r["student_id"] for r in rows]
        )
    }
    for row in rows:
        sid = row["student_id"]
        ex = existing.get(sid)
        entry = {
            "student_id": sid,
            "is_conflict": ex is not None,
            "incoming": row,
            "existing": student_row_to_dict(ex) if ex else None,
        }
        out.append(entry)
    return out


def apply_import(
    rows: list[dict],
    resolutions: dict[str, str],
) -> tuple[int, int, list[str]]:
    """
    resolutions: student_id -> 'overwrite' | 'skip'
    仅当学号已存在时适用 skip；新生默认导入。
    """
    imported = 0
    skipped = 0
    errors: list[str] = []
    for row in rows:
        sid = row["student_id"]
        exists = Student.objects.filter(student_id=sid).exists()
        res = resolutions.get(sid, "overwrite")
        if exists and res == "skip":
            skipped += 1
            continue
        if exists and res == "overwrite":
            StudentActivity.objects.filter(student_id=sid).delete()

        student, _ = Student.objects.update_or_create(
            student_id=sid,
            defaults={
                "name": row["name"],
                "department": row.get("department", ""),
                "grade": row.get("grade", ""),
                "class_name": row.get("class_name", ""),
                "academic_score": row.get("academic_score", Decimal("0")),
                "remark": row.get("remark", ""),
            },
        )
        for field, category in CATEGORY_BY_FIELD.items():
            for act_name, score in row.get(field, []):
                act = Activity.objects.filter(name=act_name, category=category).first()
                if act is None:
                    act = Activity.objects.create(
                        activity_id=_new_activity_id(),
                        name=act_name,
                        category=category,
                        score=score,
                        is_advanced=False,
                    )
                StudentActivity.objects.create(
                    student=student,
                    activity=act,
                    actual_score=score,
                    participate_date=date.today(),
                )
        imported += 1
    return imported, skipped, errors


def export_workbook(students_qs, participations_qs) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "学生成绩汇总"
    headers1 = [
        "学号",
        "姓名",
        "院系",
        "年级",
        "班级",
        "学业水平成绩",
        "身心素养总分",
        "文艺素养总分",
        "劳动素养总分",
        "创新素养总分",
        "素质总分",
        "综合分值",
        "备注",
    ]
    ws1.append(headers1)
    for s in students_qs:
        ws1.append(
            [
                s.student_id,
                s.name,
                s.department,
                s.grade,
                s.class_name,
                float(s.academic_score),
                float(getattr(s, "mind_total", 0) or 0),
                float(getattr(s, "art_total", 0) or 0),
                float(getattr(s, "labor_total", 0) or 0),
                float(getattr(s, "innovation_total", 0) or 0),
                float(s.quality_total),
                float(s.comprehensive_score),
                s.remark,
            ]
        )
    ws2 = wb.create_sheet("活动参与明细")
    headers2 = [
        "学号",
        "姓名",
        "活动名称",
        "所属类别",
        "标准分值",
        "实际得分",
        "是否为进阶",
        "参与时间",
    ]
    ws2.append(headers2)
    for p in participations_qs.select_related("student", "activity"):
        eff = p.actual_score if p.actual_score is not None else p.activity.score
        ws2.append(
            [
                p.student.student_id,
                p.student.name,
                p.activity.name,
                p.activity.category,
                float(p.activity.score),
                float(eff),
                "是" if p.activity.is_advanced else "否",
                p.participate_date.isoformat() if p.participate_date else "",
            ]
        )
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
